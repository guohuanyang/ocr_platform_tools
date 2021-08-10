# -*- coding: utf-8 -*-

# -------------------------------------------------------------------------------
# Name:         excel_manager
# Description:  
# Author:       guohuanyang
# Date:         2021/8/10
# Email         guohuanyang@datagrand.com
# -------------------------------------------------------------------------------
import openpyxl
from config.conf import TEMPLATE_NAME_DICT


class ExcelManager:

    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.Workbook()

    def create_sheets(self, sheet_name):
        self.wb.create_sheet(sheet_name)

    def write_excel(self, sheet_name, result):
        ws = self.wb.get_sheet_by_name(sheet_name)
        for row in result:
            ws.append(row)

    def save(self, filepath=None):
        self.wb.save(filepath or self.filepath)
        self.wb.close()
